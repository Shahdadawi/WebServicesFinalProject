from fastapi import APIRouter , Query
from fastapi.responses import JSONResponse
from datetime import datetime
from reportlab.pdfgen import canvas
from fastapi.responses import FileResponse
from openpyxl import Workbook

from backend.database import incident_reports

router = APIRouter()

@router.get("/by-location")
def analytics_by_location():
    pipeline = [
        {"$group": {"_id": "$incident_details.location.city", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    results = incident_reports.aggregate(pipeline)
    summary = {item["_id"]: item["count"] for item in results if item["_id"]}
    return JSONResponse(content=summary)

@router.get("/by-date")
def analytics_by_date():
    pipeline = [
        {"$project": {"date": {"$dateToString": {"format": "%Y-%m", "date": "$created_at"}}}},
        {"$group": {"_id": "$date", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}}
    ]
    results = incident_reports.aggregate(pipeline)
    summary = {item["_id"]: item["count"] for item in results if item["_id"]}
    return JSONResponse(content=summary)

@router.get("/summary")
def full_analytics():
    violations = list(incident_reports.aggregate([
        {"$unwind": "$incident_details.violation_types"},
        {"$group": {"_id": "$incident_details.violation_types", "count": {"$sum": 1}}}
    ]))
    locations = list(incident_reports.aggregate([
        {"$group": {"_id": "$incident_details.location.city", "count": {"$sum": 1}}}
    ]))
    by_date = list(incident_reports.aggregate([
        {"$project": {"date": {"$dateToString": {"format": "%Y-%m", "date": "$created_at"}}}},
        {"$group": {"_id": "$date", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}}
    ]))

    return {
        "violations": {v["_id"]: v["count"] for v in violations},
        "locations": {l["_id"]: l["count"] for l in locations if l["_id"]},
        "by_date": {d["_id"]: d["count"] for d in by_date}
    }




@router.get("/filtered")
def filtered_analytics(
    start_date: str = Query(None),
    end_date: str = Query(None),
    city: str = Query(None),
    violation_type: str = Query(None)
):
    match_stage = {}
    
    if start_date and end_date:
        match_stage["created_at"] = {
            "$gte": datetime.strptime(start_date, "%Y-%m-%d"),
            "$lte": datetime.strptime(end_date, "%Y-%m-%d")
        }
    if city:
        match_stage["incident_details.location.city"] = city
    if violation_type:
        match_stage["incident_details.violation_types"] = violation_type

    pipeline = []
    if match_stage:
        pipeline.append({"$match": match_stage})

    pipeline += [
        {"$group": {
            "_id": "$incident_details.violation_types",
            "count": {"$sum": 1}
        }},
        {"$sort": {"count": -1}}
    ]

    result = list(incident_reports.aggregate(pipeline))
    return {item["_id"]: item["count"] for item in result}


@router.get("/export/pdf")
def generate_pdf():
    file_path = "report.pdf"
    c = canvas.Canvas(file_path)
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, 800, "Human Rights Report Summary")

    c.setFont("Helvetica", 12)
    c.drawString(50, 780, f"Generated on: {datetime.now().strftime('%Y-%m-%d')}")

    y = 750

    # Violations by Type
    violations = list(incident_reports.aggregate([
        {"$unwind": "$incident_details.violation_types"},
        {"$group": {"_id": "$incident_details.violation_types", "count": {"$sum": 1}}}
    ]))

    c.drawString(50, y, "Violation Types:")
    y -= 20
    for item in violations:
        c.drawString(70, y, f"- {item['_id']}: {item['count']}")
        y -= 20

    # Violations by City
    y -= 10
    cities = list(incident_reports.aggregate([
        {"$group": {"_id": "$incident_details.location.city", "count": {"$sum": 1}}}
    ]))

    c.drawString(50, y, "Reports by City:")
    y -= 20
    for item in cities:
        if item["_id"]:
            c.drawString(70, y, f"- {item['_id']}: {item['count']}")
            y -= 20

    # Reports by Date
    y -= 10
    by_date = list(incident_reports.aggregate([
        {"$project": {"date": {"$dateToString": {"format": "%Y-%m", "date": "$created_at"}}}},
        {"$group": {"_id": "$date", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}}
    ]))

    c.drawString(50, y, "Reports Over Time:")
    y -= 20
    for item in by_date:
        c.drawString(70, y, f"- {item['_id']}: {item['count']}")
        y -= 20

    c.save()
    return FileResponse(path=file_path, filename="report.pdf", media_type="application/pdf")


@router.get("/export/excel")
def generate_excel():
    wb = Workbook()

    # Sheet 1: Violations by City
    ws1 = wb.active
    ws1.title = "By City"
    ws1.append(["City", "Count"])
    cities = list(incident_reports.aggregate([
        {"$group": {"_id": "$incident_details.location.city", "count": {"$sum": 1}}}
    ]))
    for item in cities:
        if item["_id"]:
            ws1.append([item["_id"], item["count"]])

    # Sheet 2: Violation Types
    ws2 = wb.create_sheet("By Violation")
    ws2.append(["Violation Type", "Count"])
    violations = list(incident_reports.aggregate([
        {"$unwind": "$incident_details.violation_types"},
        {"$group": {"_id": "$incident_details.violation_types", "count": {"$sum": 1}}}
    ]))
    for item in violations:
        ws2.append([item["_id"], item["count"]])

    # Sheet 3: Timeline
    ws3 = wb.create_sheet("By Date")
    ws3.append(["Date", "Count"])
    by_date = list(incident_reports.aggregate([
        {"$project": {"date": {"$dateToString": {"format": "%Y-%m", "date": "$created_at"}}}},
        {"$group": {"_id": "$date", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}}
    ]))
    for item in by_date:
        ws3.append([item["_id"], item["count"]])

    file_path = "report.xlsx"
    wb.save(file_path)
    return FileResponse(path=file_path, filename="report.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



@router.get("/geodata")
def geo_data():
    results = incident_reports.find({}, {"incident_details.location": 1, "_id": 0})
    geo = []
    for item in results:
        loc = item["incident_details"]["location"]
        coords = loc.get("coordinates", {}).get("coordinates")
        if coords:
            geo.append({
                "city": loc["city"],
                "lat": coords[1],
                "lon": coords[0]
            })
    return geo

